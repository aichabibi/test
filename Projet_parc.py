import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Charger le fichier Excel
def load_excel(excel_path):
    data = pd.read_excel(excel_path, sheet_name='Position')
    data.set_index('Symbole', inplace=True)
    return data

# Section Recherche
def search_info(data):
    st.sidebar.header("Recherche")
    symbole_options = data.index.tolist()

    # Afficher la liste déroulante avec les options filtrées
    selected_symbole = st.sidebar.selectbox("Symbole :", options=symbole_options, index=0)
    
    if selected_symbole:
        info = data.loc[selected_symbole]
        position = info['Position']
        if position.is_integer():  # Vérifie si la position est un nombre entier
            position = int(position)  # Convertit en entier
        designation = info['Désignation']
        
        # Utilisation de style HTML pour afficher les mots en rouge et en gras
        st.write(f"<span style='color:red; font-weight:bold'>Symbole:</span> {selected_symbole}<br>"
                 f"<span style='color:red; font-weight:bold'>Position:</span> {position}<br>"
                 f"<span style='color:red; font-weight:bold'>Désignation:</span> {designation}", unsafe_allow_html=True)


# Section Modification de l'emplacement
def modify_location(data, excel_path):
    st.sidebar.header("Modification de l'emplacement")
    symbole_options = data.index.tolist()

    # Utilisation d'un identifiant spécifique comme clé unique pour le st.selectbox
    selected_symbole = st.sidebar.selectbox("Symbole :", options=symbole_options, index=0, key="selectbox_symbole")

    nouvelle_position = st.sidebar.text_input("Nouvelle position :", key="nouvelle_position")

    if selected_symbole and nouvelle_position:
        data.loc[selected_symbole, 'Position'] = nouvelle_position
        
        # Enregistrement des modifications dans le DataFrame
        wb = load_workbook(filename=excel_path)
        ws = wb['Position']
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == selected_symbole:
                ws.cell(row=row, column=3).value = nouvelle_position
                break

        wb.save(excel_path)
        st.success(f"Emplacement du symbole {selected_symbole} mis à jour avec succès.")


# Section Ajout d'un nouveau symbole
def add_symbol(data, excel_path):
    st.sidebar.header("Ajout d'un nouveau symbole")
    new_symbole = st.sidebar.text_input("Nouveau symbole :", key="new_symbole")
    new_position = st.sidebar.text_input("Nouvelle position :", key="new_position")
    new_designation = st.sidebar.text_input("Nouvelle désignation :", key="new_designation")

    if st.sidebar.button("Ajouter"):
        # Ajout du nouveau symbole dans le DataFrame
        data.loc[new_symbole] = [new_position, new_designation]
        
        # Enregistrement des modifications dans le fichier Excel
        wb = load_workbook(filename=excel_path)
        ws = wb['Position']
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = new_symbole
        ws.cell(row=next_row, column=2).value = new_position
        ws.cell(row=next_row, column=3).value = new_designation
        wb.save(excel_path)
        st.success(f"Nouveau symbole {new_symbole} ajouté avec succès.")

# Section Suppression d'un symbole
def delete_symbol(data, excel_path):
    st.sidebar.header("Suppression d'un symbole")
    symbole_to_delete = st.sidebar.selectbox("Symbole à supprimer :", options=data.index.tolist(), key="symbole_to_delete")

    if st.sidebar.button("Supprimer"):
        # Suppression du symbole sélectionné du DataFrame
        data = data.drop(symbole_to_delete)
        
        # Enregistrement des modifications dans le fichier Excel
        wb = load_workbook(filename=excel_path)
        ws = wb['Position']
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == symbole_to_delete:
                ws.delete_rows(row)
                break

        wb.save(excel_path)
        st.success(f"Symbole {symbole_to_delete} supprimé avec succès.")

import os

def main():
    st.title("Gestion du matériel au PARC")
    
    # Récupération du chemin d'accès relatif de l'image
    directory = os.path.dirname(__file__)  # Récupère le répertoire du script actuel
    logo_path = os.path.join(directory, "eiffagerail240001colourrgb-002-1663689976.png")  # Joindre le répertoire avec le nom de l'image

    st.sidebar.image(logo_path, use_column_width=True)

    st.sidebar.title("Paramètres")
    excel_path = r"C:\Users\ABIBI\Documents\Projet Parc\Localisation du matériel au PARC.xlsx"
    if st.sidebar.button("Réinitialiser"):
        st.experimental_rerun()
    if excel_path is not None:
        data = load_excel(excel_path)

        search_info(data)
        modify_location(data, excel_path)
        add_symbol(data, excel_path)
        delete_symbol(data, excel_path)


if __name__ == "__main__":
    main()